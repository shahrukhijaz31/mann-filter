import os
import re
import time
import json
import requests
import threading
import platform
import subprocess
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, font
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging

# Constants
OUTPUT_DIR = "output_excels"
DYNAMIC_WAIT_TIME = 15
CHUNK_SIZE = 100  # Pages per file

# Base URL
base_url = "https://www.mann-filter.com/api/graphql/catalog-prod"

# GraphQL query
query = """
query($search:String! $currentPage:Int! $pageSize:Int! $filterBy:TYPE_OF_FILTER) {
  catalogSearch:search_crossreference_no(
    search:$search
    currentPage:$currentPage
    pageSize:$pageSize
    filterBy:$filterBy
  ) {
    items {
      product {
        name
        sku
        attributes:attributes_value {
          key
          value
          __typename
        }
        __typename
      }
      externalNumber:external_number
      manufacturer:ext_brand_name
      filterBy:aa_product_family
      __typename
    }
    pageInfo:page_info {
      currentPage:current_page
      pageSize:page_size
      totalPages:total_pages
      __typename
    }
    totalCount:total_count
    __typename
  }
}
"""


class ScraperApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("MANN-FILTER Scraper Pro")
        self.root.geometry("750x550")

        # MANN-FILTER color scheme
        self.primary_color = "#5EC619"
        self.secondary_color = "#E00A47"
        self.background_color = "#FFF000"
        self.text_color = "#333333"
        self.highlight_color = "#ffffff"

        # Configure root window
        self.root.configure(bg=self.background_color)

        # Custom fonts
        self.title_font = font.Font(family="Arial", size=16, weight="bold")
        self.header_font = font.Font(family="Arial", size=12, weight="bold")
        self.normal_font = font.Font(family="Arial", size=10)
        self.mono_font = font.Font(family="Consolas", size=9)

        # Style configuration
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure(
            '.', background=self.background_color, foreground=self.text_color)
        self.style.configure('TFrame', background=self.background_color)
        self.style.configure(
            'TLabel', background=self.background_color, font=self.normal_font)
        self.style.configure('TButton', font=self.normal_font, padding=5)
        self.style.configure('TEntry', font=self.normal_font, padding=5)
        self.style.configure(
            'Header.TLabel', font=self.header_font, foreground=self.secondary_color)
        self.style.configure(
            'Title.TLabel', font=self.title_font, foreground=self.primary_color)
        self.style.configure('Horizontal.TProgressbar',
                             troughcolor='#dddddd', background=self.primary_color)
        self.style.configure('Green.TFrame', background="#007D50")
        self.style.configure('Red.TFrame', background="#007D50")
        self.style.configure('Yellow.TFrame', background="#FFC107")

        # Create UI components
        self.create_header()
        self.create_main_container()
        self.create_input_frame()
        self.create_progress_frame()
        self.create_console_frame()
        self.create_button_frame()

        # Scraper control variables
        self.is_scraping = False
        self.should_stop = False
        self.start_time = None
        self.processed_count = 0
        self.current_product = ""

        # Configure tags for console coloring
        self.console.tag_config('error', foreground='red')
        self.console.tag_config('warning', foreground='orange')
        self.console.tag_config('success', foreground='green')
        self.console.tag_config('info', foreground='blue')
        self.console.tag_config('mann', foreground=self.primary_color)

        # Suppress unnecessary logs
        self.suppress_logs()

    def suppress_logs(self):
        """Suppress unnecessary log messages"""
        os.environ['WDM_LOG_LEVEL'] = '0'
        os.environ['WDM_PRINT_FIRST_LINE'] = 'False'

        # Suppress Selenium logs
        from selenium.webdriver.remote.remote_connection import LOGGER
        import logging
        LOGGER.setLevel(logging.WARNING)

    def create_header(self):
        header_frame = ttk.Frame(self.root, style='TFrame')
        header_frame.pack(fill=tk.X, pady=(0, 10))

        logo_frame = ttk.Frame(header_frame, style='TFrame')
        logo_frame.pack(side=tk.LEFT, padx=10, pady=5)
        ttk.Label(logo_frame, text="MANN-FILTER",
                  style='Title.TLabel').pack(side=tk.LEFT)
        ttk.Label(logo_frame, text="Product Scraper",
                  style='Header.TLabel').pack(side=tk.LEFT, padx=(5, 0))

        status_frame = ttk.Frame(header_frame, style='TFrame')
        status_frame.pack(side=tk.RIGHT, padx=10, pady=5)
        self.status_frame = ttk.Frame(
            status_frame, style='Red.TFrame', height=20, width=20)
        self.status_frame.pack(side=tk.LEFT, padx=5)
        self.status_label = ttk.Label(
            status_frame, text="Ready", foreground=self.text_color)
        self.status_label.pack(side=tk.LEFT)

    def create_main_container(self):
        self.main_canvas = tk.Canvas(
            self.root, borderwidth=0, background=self.background_color, highlightthickness=0)
        self.main_frame = ttk.Frame(self.main_canvas, style='TFrame')
        self.vsb = tk.Scrollbar(
            self.root, orient="vertical", command=self.main_canvas.yview)
        self.main_canvas.configure(yscrollcommand=self.vsb.set)
        self.vsb.pack(side="right", fill="y")
        self.main_canvas.pack(side="left", fill="both", expand=True)
        self.main_canvas.create_window(
            (4, 4), window=self.main_frame, anchor="nw")
        self.main_frame.bind("<Configure>", lambda e: self.main_canvas.configure(
            scrollregion=self.main_canvas.bbox("all")))

    def create_input_frame(self):
        input_frame = ttk.LabelFrame(
            self.main_frame, text="Search Parameters", padding=10, style='TFrame')
        input_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        input_frame.grid_columnconfigure(1, weight=1)

        # Search term range
        ttk.Label(input_frame, text="Search Term Range:").grid(
            row=0, column=0, sticky=tk.W, padx=5, pady=5)
        search_range_frame = ttk.Frame(input_frame, style='TFrame')
        search_range_frame.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=5)
        self.search_start_entry = ttk.Entry(search_range_frame, width=10)
        self.search_start_entry.pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(search_range_frame, text="to").pack(side=tk.LEFT, padx=5)
        self.search_end_entry = ttk.Entry(search_range_frame, width=10)
        self.search_end_entry.pack(side=tk.LEFT)

        # Page range
        ttk.Label(input_frame, text="Page Range (1-667):").grid(row=1,
                                                                column=0, sticky=tk.W, padx=5, pady=5)
        page_range_frame = ttk.Frame(input_frame, style='TFrame')
        page_range_frame.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        self.page_start_spin = ttk.Spinbox(
            page_range_frame, from_=1, to=667, width=5)
        self.page_start_spin.pack(side=tk.LEFT, padx=(0, 5))
        self.page_start_spin.set(1)
        ttk.Label(page_range_frame, text="to").pack(side=tk.LEFT, padx=5)
        self.page_end_spin = ttk.Spinbox(
            page_range_frame, from_=1, to=667, width=5)
        self.page_end_spin.pack(side=tk.LEFT)
        self.page_end_spin.set(5)

        # Output filename base
        ttk.Label(input_frame, text="Output Filename Base:").grid(
            row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.filename_entry = ttk.Entry(input_frame, width=50)
        default_filename = "mann_products"
        self.filename_entry.insert(0, default_filename)
        self.filename_entry.grid(row=2, column=1, padx=5, pady=5, sticky=tk.EW)

    def create_progress_frame(self):
        self.progress_frame = ttk.LabelFrame(
            self.main_frame, text="Scraping Progress", padding=10, style='TFrame')
        self.progress_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        # Page progress
        ttk.Label(self.progress_frame, text="Page Progress:").pack(anchor=tk.W)
        self.page_progress = ttk.Progressbar(self.progress_frame, orient=tk.HORIZONTAL,
                                             length=300, mode='determinate', style='Horizontal.TProgressbar')
        self.page_progress.pack(fill=tk.X, pady=(0, 5))
        self.page_label = ttk.Label(self.progress_frame, text="Pages: 0/0")
        self.page_label.pack(anchor=tk.W)

        # Current product
        ttk.Label(self.progress_frame,
                  text="Current Product:").pack(anchor=tk.W)
        self.current_product_label = ttk.Label(self.progress_frame, text="",
                                               font=('Arial', 9, 'italic'), foreground="#2c3e50")
        self.current_product_label.pack(anchor=tk.W, fill=tk.X, padx=5)

        # Product progress
        ttk.Label(self.progress_frame,
                  text="Product Progress:").pack(anchor=tk.W)
        self.product_progress = ttk.Progressbar(self.progress_frame, orient=tk.HORIZONTAL,
                                                length=300, mode='determinate', style='Horizontal.TProgressbar')
        self.product_progress.pack(fill=tk.X, pady=(0, 5))
        self.product_label = ttk.Label(self.progress_frame, text="Products: 0")
        self.product_label.pack(anchor=tk.W)

        # Stats
        stats_frame = ttk.Frame(self.progress_frame, style='TFrame')
        stats_frame.pack(fill=tk.X, pady=(5, 0))
        self.time_label = ttk.Label(stats_frame, text="Elapsed: 00:00:00")
        self.time_label.pack(side=tk.LEFT)
        self.rate_label = ttk.Label(stats_frame, text="Rate: 0 products/min")
        self.rate_label.pack(side=tk.RIGHT)

    def create_console_frame(self):
        console_frame = ttk.LabelFrame(
            self.main_frame, text="Console Output", padding=6, style='TFrame')
        console_frame.pack(fill=tk.BOTH, expand=True, padx=5)
        self.console = scrolledtext.ScrolledText(console_frame, height=12, state='disabled', wrap=tk.WORD,
                                                 font=self.mono_font, background='#ffffff', foreground=self.text_color)
        self.console.pack(fill=tk.BOTH, expand=True)

    def create_button_frame(self):
        button_frame = ttk.Frame(self.main_frame, style='TFrame')
        button_frame.pack(fill=tk.X, pady=(10, 5), padx=5)
        self.style.configure(
            'Primary.TButton', foreground=self.highlight_color, background=self.primary_color)
        self.style.configure(
            'Secondary.TButton', foreground=self.highlight_color, background=self.secondary_color)

        self.start_btn = ttk.Button(
            button_frame, text="Start Scraping", command=self.start_scraping, style='Primary.TButton')
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.cancel_btn = ttk.Button(
            button_frame, text="Cancel", command=self.cancel_scraping, state='disabled', style='Secondary.TButton')
        self.cancel_btn.pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Help", command=self.show_help,
                   style='Secondary.TButton').pack(side=tk.RIGHT, padx=5)

    def log_message(self, message, level='info'):
        """Add message to console with timestamp and colored by level"""
        self.console.config(state='normal')
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.console.insert(tk.END, f"[{timestamp}] ", 'mann')
        self.console.insert(tk.END, f"{message}\n", level)
        self.console.see(tk.END)
        self.console.config(state='disabled')
        self.root.update_idletasks()

    def update_page_progress(self, current, total):
        """Update page progress bar and label"""
        self.page_progress['maximum'] = total
        self.page_progress['value'] = current
        self.page_label.config(text=f"Pages: {current}/{total}")
        self.update_status()
        self.root.update_idletasks()

    def update_product_progress(self, count):
        """Update product progress bar and label"""
        self.product_progress['value'] = count
        self.product_label.config(text=f"Products: {count}")
        self.processed_count = count
        self.update_status()
        self.update_stats()
        self.root.update_idletasks()

    def update_current_product(self, product_name):
        """Update the current product being processed"""
        self.current_product = product_name
        display_name = product_name[:80] + \
            ("..." if len(product_name) > 80 else "")
        self.current_product_label.config(text=display_name)
        self.root.update_idletasks()

    def update_status(self):
        """Update the status indicator"""
        if self.is_scraping:
            self.status_frame.config(style='Green.TFrame')
            self.status_label.config(text="Working", foreground="#27ae60")
        else:
            self.status_frame.config(style='Red.TFrame')
            self.status_label.config(text="Ready", foreground="#7f8c8d")

    def update_stats(self):
        """Update time and rate statistics"""
        if not self.start_time:
            return

        elapsed = datetime.now() - self.start_time
        elapsed_str = str(elapsed).split('.')[0]  # Remove microseconds

        # Calculate rate (products per minute)
        minutes = elapsed.total_seconds() / 60
        rate = self.processed_count / minutes if minutes > 0 else 0

        self.time_label.config(text=f"Elapsed: {elapsed_str}")
        self.rate_label.config(text=f"Rate: {rate:.1f} products/min")

    def get_output_filename(self, search_term, start_page, end_page):
        """Generate clean filename without timestamp"""
        return f"{self.filename_entry.get().strip()}_searchterm_{search_term}_{start_page}-{end_page}.xlsx"

    def get_last_completed_page(self, search_term):
        """Find where we left off for a search term"""
        max_page = 0
        pattern = re.compile(
            rf"{re.escape(self.filename_entry.get().strip())}_searchterm_{search_term}_(\d+)-(\d+).xlsx")

        if not os.path.exists(OUTPUT_DIR):
            return 0

        for filename in os.listdir(OUTPUT_DIR):
            match = pattern.match(filename)
            if match:
                start = int(match.group(1))
                end = int(match.group(2))
                if end > max_page:
                    max_page = end
        return max_page

    def start_scraping(self):
        """Start scraping in a separate thread"""
        try:
            search_start = self.search_start_entry.get().strip()
            search_end = self.search_end_entry.get().strip()
            page_start = int(self.page_start_spin.get())
            page_end = int(self.page_end_spin.get())
            filename_base = self.filename_entry.get().strip()

            # Validate inputs
            if not all([search_start, search_end, filename_base]):
                messagebox.showerror(
                    "Error", "Please fill all required fields")
                return

            try:
                search_start = int(search_start)
                search_end = int(search_end)
            except ValueError:
                messagebox.showerror(
                    "Error", "Search term range must be numbers")
                return

            if search_start > search_end:
                messagebox.showerror(
                    "Error", "Start search term must be <= end search term")
                return

            if page_start > page_end:
                messagebox.showerror("Error", "Start page must be <= end page")
                return

            self.is_scraping = True
            self.should_stop = False
            self.start_btn.config(state='disabled')
            self.cancel_btn.config(state='normal')
            self.start_time = datetime.now()
            self.processed_count = 0

            # Reset UI elements
            self.page_progress['value'] = 0
            self.product_progress['value'] = 0
            self.product_progress['maximum'] = 0
            self.current_product_label.config(text="")
            self.console.config(state='normal')
            self.console.delete(1.0, tk.END)
            self.console.config(state='disabled')

            # Start scraping thread
            thread = threading.Thread(
                target=self.run_scraper,
                args=(search_start, search_end, page_start, page_end),
                daemon=True
            )
            thread.start()

            # Start stats update timer
            self.update_stats_timer()

        except Exception as e:
            self.log_message(f"Error starting scrape: {str(e)}", 'error')
            messagebox.showerror(
                "Error", f"Failed to start scraping: {str(e)}")

    def update_stats_timer(self):
        """Periodically update stats while scraping"""
        if self.is_scraping:
            self.update_stats()
            self.root.after(1000, self.update_stats_timer)

    def run_scraper(self, search_start, search_end, page_start, page_end):
        """Main scraping function with file-based progress tracking"""
        try:
            os.makedirs(OUTPUT_DIR, exist_ok=True)

            for search_term in range(search_start, search_end + 1):
                if self.should_stop:
                    break

                # Determine resume point for this search term
                last_completed_page = self.get_last_completed_page(search_term)
                current_page = max(
                    page_start, last_completed_page + 1) if last_completed_page > 0 else page_start

                self.log_message(
                    f"\nProcessing search term: {search_term}", 'info')
                total_pages = page_end
                self.update_page_progress(current_page, total_pages)

                while current_page <= page_end:
                    if self.should_stop:
                        break

                    # Calculate current chunk
                    chunk_start = current_page
                    chunk_end = min(current_page + CHUNK_SIZE - 1, page_end)

                    # Create filename for this chunk
                    filename = self.get_output_filename(
                        search_term, chunk_start, chunk_end)
                    filepath = os.path.join(OUTPUT_DIR, filename)

                    # Skip if this chunk already exists
                    if os.path.exists(filepath):
                        self.log_message(
                            f"Skipping existing chunk: {filename}", 'info')
                        current_page = chunk_end + 1
                        continue

                    self.log_message(
                        f"Processing pages {chunk_start} to {chunk_end}...", 'info')

                    # Process all pages in this chunk
                    chunk_products = []
                    for page in range(chunk_start, chunk_end + 1):
                        if self.should_stop:
                            break

                        self.log_message(f"Fetching page {page}...")
                        self.update_page_progress(page, total_pages)

                        try:
                            # Make API request
                            response = requests.get(
                                base_url,
                                params={
                                    "query": query,
                                    "variables": json.dumps({
                                        "search": str(search_term),
                                        "currentPage": page,
                                        "pageSize": 15,
                                        "filterBy": "ALL_FILTER"
                                    })
                                },
                                timeout=30
                            )
                            response.raise_for_status()
                            data = response.json()

                            if 'errors' in data:
                                self.log_message(
                                    f"API Error: {data['errors']}", 'error')
                                continue

                            catalog_data = data['data']['catalogSearch']

                            # Update total pages if first page
                            if page == chunk_start:
                                total_items = catalog_data.get('totalCount', 0)
                                self.log_message(
                                    f"Found {total_items} total items", 'info')
                                self.product_progress['maximum'] = total_items

                            # Extract product references
                            page_items = []
                            for item in catalog_data['items']:
                                if item.get('product'):
                                    product_ref = {
                                        'id': len(chunk_products) + len(page_items) + 1,
                                        'external_number': item.get('externalNumber', ''),
                                        'manufacturer': item.get('manufacturer', ''),
                                        'mann_filter': item['product'].get('sku', ''),
                                        'status': extract_product_status(item['product'].get('attributes', [])),
                                        'filter_type': item.get('filterBy', ''),
                                        'url': f"https://www.mann-filter.com/tr-tr/katalog/arama-sonuclar%C4%B1/urun.html/{item['product'].get('sku', '').lower()}.html"
                                    }
                                    page_items.append(product_ref)

                            self.log_message(
                                f"Processing {len(page_items)} products...")
                            for product_ref in page_items:
                                if self.should_stop:
                                    break

                                # Update current product display
                                self.update_current_product(
                                    product_ref['mann_filter'])

                                # Process product URL
                                product_data = process_url(
                                    product_ref['id'],
                                    product_ref['external_number'],
                                    product_ref['manufacturer'],
                                    product_ref['mann_filter'],
                                    product_ref['status'],
                                    product_ref['filter_type'],
                                    product_ref['url']
                                )

                                if product_data:
                                    chunk_products.append(product_data)
                                    self.processed_count += 1
                                    self.update_product_progress(
                                        self.processed_count)

                            time.sleep(1)  # Be polite to the server

                        except Exception as e:
                            self.log_message(
                                f"Error processing page {page}: {str(e)}", 'error')
                            time.sleep(5)  # Wait before retrying
                            continue

                    # Save chunk after processing
                    if chunk_products and not self.should_stop:
                        success = create_combined_excel(
                            chunk_products, filename)
                        if success:
                            self.log_message(
                                f"Saved {len(chunk_products)} products to {filename}", 'success')
                        else:
                            self.log_message(
                                f"Failed to save {filename}", 'error')

                    # Move to next chunk
                    current_page = chunk_end + 1

                if not self.should_stop:
                    self.log_message(
                        f"Completed search term {search_term}", 'success')

            if self.should_stop:
                self.log_message("\nScraping cancelled by user", 'warning')
                messagebox.showinfo("Cancelled", "Scraping was cancelled")
            else:
                self.log_message("\nAll search terms processed", 'success')
                messagebox.showinfo(
                    "Complete", "Scraping completed for all search terms!")

        except Exception as e:
            self.log_message(f"\nFatal error: {str(e)}", 'error')
            messagebox.showerror("Error", f"A fatal error occurred: {str(e)}")
        finally:
            self.is_scraping = False
            self.root.after(0, lambda: self.start_btn.config(state='normal'))
            self.root.after(
                0, lambda: self.cancel_btn.config(state='disabled'))
            self.log_message("\nScraping completed", 'info')
            self.update_status()

    def cancel_scraping(self):
        """Cancel the current scraping operation"""
        if self.is_scraping:
            self.should_stop = True
            self.log_message("Cancellation requested...", 'warning')
            self.cancel_btn.config(state='disabled')

    def show_help(self):
        """Show help information"""
        help_text = """
            MANN-FILTER Scraper Pro Help

            1. Enter a search term range (e.g., from 100 to 200)
            2. Set the page range to scrape (1-667)
            3. Specify an output filename base
            4. Click 'Start Scraping' to begin

            The scraper will:
            - Search for matching products on MANN-FILTER.com
            - Extract detailed product information
            - Save results to Excel files (one per 100 pages per search term)
            - Automatically resume from last position if stopped

            Output Files:
            - Format: [filename]_searchterm_[N]_[start]-[end].xlsx
            - Example: mann_products_searchterm_100_1-100.xlsx
            """
        messagebox.showinfo("Help", help_text.strip())

    def run(self):
        """Run the application"""
        self.root.mainloop()


def extract_product_status(attributes):
    """Extract product status from attributes"""
    if not attributes:
        return "Unknown"
    for attr in attributes:
        if attr.get('key') == 'product_status_aa':
            return attr.get('value', 'Unknown')
    return "Unknown"


def get_driver():
    """Create and configure a Selenium WebDriver"""
    # Clean up any existing drivers
    system = platform.system().lower()
    if 'windows' in system:
        subprocess.run(['taskkill', '/F', '/IM', 'chromedriver.exe', '/T'],
                       stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL)
    elif 'linux' in system or 'darwin' in system:
        subprocess.run(['pkill', '-f', 'chromedriver'],
                       stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL)

    # Configure Chrome options
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--log-level=3')  # Suppress console logs
    options.add_experimental_option(
        'excludeSwitches', ['enable-logging'])  # Disable logging
    options.add_argument('--disable-gcm')  # Disable Google Cloud Messaging

    return webdriver.Chrome(options=options)


def get_page_content(url, driver):
    """Get page content using Selenium to handle dynamic content"""
    try:
        driver.get(url)

        # Wait for the product name to load
        WebDriverWait(driver, DYNAMIC_WAIT_TIME).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, ".cmp-product__title-name"))
        )

        # Wait for either dimensions, applications, or OE numbers section
        WebDriverWait(driver, DYNAMIC_WAIT_TIME).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, ".cmp-table, .cmp-accordion"))
        )

        # Scroll to ensure all dynamic content loads
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)

        return driver.page_source
    except Exception as e:
        print(f"Error loading {url}: {e}")
        return None


def extract_product_data(soup, id, external_number, manufacturer, mann_filter, status, filter_type, url):
    # Helper functions
    def clean_single_year(year_str):
        if not year_str:
            return ""
        year_str = re.sub(r'\D', '', year_str)
        if len(year_str) == 2:
            year_num = int(year_str)
            return f"19{year_str}" if year_num >= 30 and year_num < 100 else f"20{year_str}"
        if len(year_str) == 1:
            return f"190{year_str}" if int(year_str) >= 5 else f"200{year_str}"
        return year_str if len(year_str) == 4 else ""

    def clean_year(year_str):
        if not year_str:
            return ""
        year_str = re.sub(r'[^\d/-→]', '', year_str)
        if '-' in year_str:
            parts = year_str.split('-')
            return f"{clean_single_year(parts[0])}-{clean_single_year(parts[-1])}"
        if '→' in year_str:
            parts = year_str.split('→')
            return f"{clean_single_year(parts[0])}-{clean_single_year(parts[-1])}"
        return clean_single_year(year_str)

    def safe_text(element):
        """Safely extract text from an element that might be None"""
        return element.get_text(strip=True) if element else ''

    # Extract product name
    mann_filter_elem = soup.find('span', class_=lambda c: c and (
        'cmp-product__title-name' in c or
        'productFullDetail__productName' in c
    ))
    mann_filter = mann_filter_elem.get_text(
        strip=True) if mann_filter_elem else mann_filter

    data = {
        'id': id,
        'external_number': external_number,
        'manufacturer': manufacturer,
        'mann_filter': mann_filter,
        'status': status,
        'filter_type': filter_type,
        'url': url,
        'dimensions': {},
        'applications': {},
        'oem_numbers': []
    }

    # 1. Extract Dimensions
    try:
        dimensions_section = (
            soup.find("div", id="dimensions") or
            soup.find("div", id="boyutlar") or
            soup.find(['h2', 'h3'], string=lambda t: t and
                      ('dimension' in t.lower() or
                       'boyut' in t.lower()))
        )

        if dimensions_section:
            if dimensions_section.name in ['h2', 'h3']:
                accordion = dimensions_section.find_parent(
                    "div", class_="cmp-accordion__item")
                if accordion:
                    dimensions_section = accordion.find(
                        "div", class_="cmp-accordion__panel")

            tables = []
            if dimensions_section:
                if dimensions_section.find("table"):
                    tables = dimensions_section.find_all("table")
                else:
                    table_containers = dimensions_section.find_all(
                        class_="cmp-table__container")
                    for container in table_containers:
                        table = container.find("table")
                        if table:
                            tables.append(table)

            for table in tables:
                for row in table.find_all("tr"):
                    if not row.find("td"):
                        continue
                    cells = row.find_all("td")
                    if len(cells) == 2:
                        key = safe_text(cells[0])
                        value = safe_text(cells[1])
                        if key and value:
                            data['dimensions'][key] = value
                    elif len(cells) > 2:
                        key_cell = None
                        value_cell = None
                        for i, cell in enumerate(cells):
                            text = safe_text(cell).strip()
                            if len(text) == 1 and text.isalpha():
                                key_cell = cell
                                if i + 1 < len(cells):
                                    value_cell = cells[i + 1]
                                break
                        if key_cell and value_cell:
                            key = safe_text(key_cell)
                            value = safe_text(value_cell)
                            if key and value:
                                data['dimensions'][key] = value
    except Exception as e:
        print(f"Error extracting dimensions: {e}")

    # 2. Extract Applications
    try:
        applications_section = (
            soup.find("div", id="applications") or
            soup.find("div", id="uygulamalar") or
            soup.find("div", id="araclar") or
            soup.find("div", id="vehicles") or
            soup.find(lambda tag: tag.name in ['div', 'section'] and
                      any(keyword in safe_text(tag).lower()
                      for keyword in ['application', 'uygulama', 'araç', 'vehicle']))
        )

        if applications_section:
            brand_items = applications_section.find_all(
                "div", class_=lambda c: c and 'accordion__item' in c)
            for brand_item in brand_items:
                brand_name_elem = brand_item.find(
                    class_=lambda c: c and 'accordion__title' in c)
                brand_name = safe_text(brand_name_elem)
                if not brand_name:
                    continue

                # Clean brand name
                # brand_name = re.sub(r'\(.*?\)', '', brand_name).strip()

                model_accordion = brand_item.find(
                    class_=lambda c: c and 'accordion__nested' in c)

                if model_accordion:
                    model_items = model_accordion.find_all(
                        "div", class_=lambda c: c and 'accordion__item' in c)
                    for model_item in model_items:
                        model_name_elem = model_item.find(
                            class_=lambda c: c and 'accordion__title' in c)
                        model_name = safe_text(model_name_elem)
                        if not model_name:
                            continue

                        # Clean model name
                        # model_name = re.sub(r'\(.*?\)', '', model_name).strip()

                        # Find all tables for this model
                        tables = model_item.find_all("table")
                        if tables:
                            # Initialize brand entry
                            if brand_name not in data['applications']:
                                data['applications'][brand_name] = {}

                            # Initialize model entry
                            if model_name not in data['applications'][brand_name]:
                                data['applications'][brand_name][model_name] = []

                            # Parse each table and extend applications
                            for table in tables:
                                # Extract applications from table
                                applications = []
                                # Header-based approach
                                headers = [safe_text(th).replace('\n', ' ').replace('\t', ' ').strip()
                                           for th in table.find_all("th") if safe_text(th)]

                                # Skip header row
                                for row in table.find_all("tr")[1:]:
                                    cells = row.find_all("td")
                                    if len(cells) >= len(headers):
                                        app = {
                                            headers[i]: safe_text(cells[i]).replace(
                                                '\n', ' ').replace('\t', ' ').strip()
                                            for i in range(min(len(headers), len(cells)))
                                        }
                                        if any(app.values()):
                                            applications.append(app)

                                if applications:
                                    # Clean year format for each application
                                    for app in applications:
                                        year_key = 'Üretim yılı' if 'Üretim yılı' in app else 'Year of Manufacture'
                                        if year_key in app:
                                            app[year_key] = clean_year(
                                                app[year_key])
                                    data['applications'][brand_name][model_name].extend(
                                        applications)

    except Exception as e:
        print(f"Error extracting applications: {e}")

    # 3. Extract OEM Numbers
    try:
        oem_section = (
            soup.find("div", id="oeNumbers") or
            soup.find("div", id="oemNumbers") or
            soup.find(lambda t: 'oem' in t.get('id', '').lower() or
                      'numara' in t.get('id', '').lower()) or
            soup.find(['h2', 'h3'], string=lambda t: t and
                      ('oem' in t.lower() or
                      'orijinal' in t.lower() or
                       'numara' in t.lower()))
        )

        if oem_section:
            if oem_section.name in ['h2', 'h3']:
                accordion = oem_section.find_parent(
                    "div", class_="cmp-accordion__item")
                if accordion:
                    oem_section = accordion.find(
                        "div", class_="cmp-accordion__panel")

            nested_accordion = oem_section.find(
                "div", class_="cmp-accordion cmp-accordion__nested")
            if nested_accordion:
                manufacturer_items = nested_accordion.find_all(
                    "div", class_="cmp-accordion__item")

                for item in manufacturer_items:
                    title_elem = item.find(
                        "span", class_="cmp-accordion__title")
                    manufacturer_name = safe_text(
                        title_elem).strip() if title_elem else ""

                    panel = item.find("div", class_="cmp-accordion__panel")
                    if panel:
                        numbers = []
                        ul_list = panel.find_all("ul")
                        for ul in ul_list:
                            lis = ul.find_all("li")
                            for li in lis:
                                number = safe_text(li).strip()
                                if number and len(number) >= 3:
                                    numbers.append(number)

                        if numbers:
                            seen = set()
                            unique_numbers = []
                            for num in numbers:
                                if num not in seen:
                                    seen.add(num)
                                    unique_numbers.append(num)

                            data['oem_numbers'].append({
                                'manufacturer': manufacturer_name,
                                'number': ', '.join(unique_numbers)
                            })
    except Exception as e:
        print(f"Error extracting OEM numbers: {str(e)}")

    return data


def create_combined_excel(products, filename):
    """Create Excel file with all products' data"""
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        filepath = os.path.join(OUTPUT_DIR, filename)

        # Field mapping for Turkish to English
        field_mapping = {
            "Model Tipi": "Model Type",
            "Filtre Tipi": "Filter Type",
            "Motor Kodu": "Engine Code",
            "Üretim yılı": "Year of Manufacture"
        }

        # Create new workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create worksheets
        sheets = {
            "Cross References": wb.create_sheet("Cross References"),
            "Dimensions": wb.create_sheet("Dimensions"),
            "Vehicles": wb.create_sheet("Vehicles"),
            "OEM Numbers": wb.create_sheet("OEM Numbers")
        }

        # Add headers
        sheets["Cross References"].append(["ID", "External Number", "Manufacturer", "MANN-FILTER",
                                           "Status", "Filter Type", "URL"])
        sheets["Dimensions"].append(
            ["ID", "Mann Code", "Dimension Name", "Dimension Value"])
        sheets["Vehicles"].append(["ID", "Mann Code", "Brand", "Model", "Model Type",
                                   "Filter Type", "Engine Code", "ccm", "kW", "HP", "Year of Manufacture"])
        sheets["OEM Numbers"].append(
            ["ID", "Mann Code", "Manufacturer", "OEM Number"])

        # Add data to sheets
        for product in products:
            if not isinstance(product, dict):
                continue

            # Cross References
            sheets["Cross References"].append([
                product.get('id', ''),
                product.get('external_number', ''),
                product.get('manufacturer', ''),
                product.get('mann_filter', ''),
                product.get('status', ''),
                product.get('filter_type', ''),
                product.get('url', '')
            ])

            # Dimensions
            dimensions = product.get('dimensions', {})
            if isinstance(dimensions, dict):
                for dim_name, dim_value in dimensions.items():
                    sheets["Dimensions"].append([
                        product.get('id', ''),
                        product.get('mann_filter', ''),
                        dim_name,
                        dim_value
                    ])

            # Vehicles/Applications
            applications = product.get('applications', {})
            if isinstance(applications, dict):
                for brand_name, models in applications.items():
                    if not isinstance(models, dict):
                        continue

                    for model_name, app_list in models.items():
                        if not isinstance(app_list, list):
                            continue

                        for app in app_list:
                            if not isinstance(app, dict):
                                continue

                            # Map Turkish fields to English
                            mapped_app = {}
                            for key, value in app.items():
                                english_key = field_mapping.get(key, key)
                                mapped_app[english_key] = value

                            sheets["Vehicles"].append([
                                product.get('id', ''),
                                product.get('mann_filter', ''),
                                brand_name,
                                model_name,
                                mapped_app.get('Model Type', ''),
                                mapped_app.get('Filter Type', ''),
                                mapped_app.get('Engine Code', ''),
                                mapped_app.get('ccm', ''),
                                mapped_app.get('kW', ''),
                                mapped_app.get('HP', ''),
                                mapped_app.get('Year of Manufacture', '')
                            ])

            # OEM Numbers
            oem_numbers = product.get('oem_numbers', [])
            if isinstance(oem_numbers, list):
                for oem_entry in oem_numbers:
                    if not isinstance(oem_entry, dict):
                        continue

                    manufacturer = oem_entry.get('manufacturer', '')
                    numbers_str = oem_entry.get('number', '')
                    numbers = [num.strip()
                               for num in numbers_str.split(',') if num.strip()]

                    for number in numbers:
                        sheets["OEM Numbers"].append([
                            product.get('id', ''),
                            product.get('mann_filter', ''),
                            manufacturer,
                            number
                        ])

        # Auto-adjust column widths for all sheets
        for sheet_name, sheet in sheets.items():
            # Only adjust columns that have data (skip empty sheets)
            if sheet.max_row > 1:
                for col_idx in range(1, sheet.max_column + 1):
                    max_length = 0
                    # Get header for this column
                    header = sheet.cell(row=1, column=col_idx).value
                    header_length = len(str(header)) if header else 0

                    # Check all cells in the column (skip header row)
                    for row_idx in range(2, sheet.max_row + 1):
                        cell_value = sheet.cell(
                            row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            if cell_length > max_length:
                                max_length = cell_length

                    # Use header length if no data rows
                    if max_length == 0:
                        max_length = header_length
                    else:
                        # Ensure header is considered
                        max_length = max(max_length, header_length)

                    # Calculate adjusted width
                    adjusted_width = (max_length + 2) * 1.2
                    col_letter = get_column_letter(col_idx)
                    sheet.column_dimensions[col_letter].width = adjusted_width

        # Save file
        wb.save(filepath)
        return True

    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        return False


def process_url(id, external_number, manufacturer, mann_filter, status, filter_type, url):
    """Process a single product URL and return its data"""
    driver = get_driver()
    try:
        html_content = get_page_content(url, driver)
        if not html_content:
            return None

        soup = BeautifulSoup(html_content, 'html.parser')
        product_data = extract_product_data(
            soup, id, external_number, manufacturer,
            mann_filter, status, filter_type, url
        )
        return product_data

    except Exception as e:
        print(f"Error processing URL {url}: {str(e)}")
        return None
    finally:
        driver.quit()


if __name__ == "__main__":
    app = ScraperApp()
    app.run()
